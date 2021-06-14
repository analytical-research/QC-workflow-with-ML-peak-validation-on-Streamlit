# import libraries
import crop_UVCAD_plots
import math
import numbers
import numpy as np
import openpyxl
import os
import pandas as pd
import streamlit as st

from collections import namedtuple
from detectron2 import model_zoo
from detectron2.config import get_cfg
from detectron2.data import MetadataCatalog
from detectron2.engine import DefaultPredictor
from detectron2.utils.logger import setup_logger
from openpyxl import load_workbook

setup_logger()
platesize = 88  # Number of samples par plate.
replicates = ['rep1', 'rep2', 'rep3']
barcode_file_disk = 'LAB134008-PC8'
# Strings that will be embedded on the result excel sheets.
weeklyanalysis_dataloader_path = 'assay/analytical/mosaicQC'
chemdiv_dataloader_path = 'partner_data/eMolecules/eMol_QA'
chemdiv_comment = 'eMolecules QA Result DataLoader Input'
# The information used for concentration calculation and peak prediction.
thresholds_info_path = r"C:\Users\ares968\Desktop\Testing data\Data Process Templates\thresholds_for_Mnovadataprocess.xlsx"
model_path = r"C:\Users\ares968\Desktop\Testing data\Data Process Templates\Detectron2\models\output\model_final.pth"
# Location where data and the barcode files are saved.
data_folders_path = r"Y:\Data"
barcode_folder_path = r"Y:\Data\2021"


# QC data process
class main():
    def __init__(self):
        super().__init__()
        st.title('QC Data Process')
        form = st.form(key='my_form')
        self.sample_type = form.selectbox('Sample Type:', ('---', 'Weekly Analysis', 'Chem Div', 'QC Request'))
        self.year = form.text_input('Year (####):')
        self.week = form.text_input('Week (##):')
        self.chemist = form.text_input('Chemist (Aaaa):')
        submit = form.form_submit_button(label='Submit')
        if submit:
            if self.sample_type == '---' or self.sample_type == '':
                st.write('Select sample type.')
            elif self.year == '':
                st.write('Enter the year.')
            elif self.week == '':
                st.write('Enter the week.')
            else:
                paths = get_barcode_and_data_paths(self.sample_type, self.year, self.week, self.chemist)
                sampleid = os.path.basename(paths.barcode)
                self.sampleid = sampleid.split('.')[0]
                self.plateids = make_list_of_plateids(paths.barcode)
                data_process(path_to_mnovadata=paths.data, path_to_barcode=paths.barcode,
                             sampleid=self.sampleid, plateids=self.plateids)


def get_barcode_and_data_paths(sample_type, year, week, chemist):
    """
    Get the paths to the barcode file and data folder.
    :param sample_type: Weekly Analysis, and Chem Div, or QC request.
    :param year:
    :param week:
    :param chemist:
    :return: namedtuple. Each element can be called as return.barcode or return.mnova.
    """
    barcode_path = barcode_folder_path + '\\' + year + r'-barcodefile\\' + sample_type
    data_folder_path = data_folders_path + '\\' + year + '\\' + sample_type
    if sample_type == 'Weekly Analysis':
        barcode_path = barcode_path + '\\WK' + week + r'.xlsx'
        data_folder_path = data_folder_path + '\\WK' + week + '\\WK' + week + r' DataAnalysis' + '\\WK' + week + r' Mnova'
    elif sample_type == 'Chem Div':
        barcode_path = barcode_path + '\\CD' + week + r'.xlsx'
        data_folder_path = data_folder_path + '\\CD' + week + '\\CD' + week + r' DataAnalysis' + '\\CD' + week + r' Mnova'
    elif sample_type == 'QC Request':
        barcode_path = barcode_path + '\\' + chemist + '\\WK' + week + chemist + r'.xlsx'
        data_folder_path = data_folder_path + '\\' + chemist + '\\WK' + week + chemist + '\\WK' + week + chemist + r' DataAnalysis' + '\\WK' + week + chemist + r' Mnova'
    else:
        st.write('Input is not right.')
    path = namedtuple('path', ['barcode', 'data'])
    paths = path(barcode=barcode_path, data=data_folder_path)
    return paths


def make_list_of_plateids(path_to_barcode):
    """
    Make the list of plate IDs from the barcode file assuming that one plate contains maximum of 88 samples (platesize).

    :param path_to_barcode: path to the barcode file (string).
    :return: list of plate IDs.
    """
    # Read the barcode excel file.
    labsol_barcode = pd.read_excel(path_to_barcode, sheet_name='Tube Info RL', usecols=[*range(0, 15)], header=0)
    # number of samples in the sample id.
    num_samples = len(labsol_barcode)
    # Number of plates in the sample id.
    num_plates = math.ceil(num_samples / platesize)
    plateids_list = []
    # No plate id if there is only one plate.
    if num_plates == 1:
        plateids_list = ['']
    else:
        for num_plate in range(num_plates):
            # Use ASCII to convert the number of plate to plate ID.
            ascii_num = num_plate + 97
            plateid = chr(ascii_num)
            plateids_list.append(plateid)
    return plateids_list


def check_barcodefile_sheet_name(path_to_barcode):
    """
    The sheet name has to be "Tube Info RL". If not, it sends the message.
    :param path_to_barcode:  path to the barcode file (string).
    :return: Boolean
    """
    # Get the sheet names of the excel file.
    wb = load_workbook(path_to_barcode, read_only=True, keep_links=False)
    sheet_names = wb.sheetnames
    # Check if the barcode sheet is named as "Tube Info RL". If not, False.
    if "Tube Info RL" not in sheet_names:
        st.write("Error", "The barcode sheet on the excel file has to be named as Tube Info RL.")
        correct_sheet_name = False
    else:
        correct_sheet_name = True
    return correct_sheet_name


def get_std_info(path_to_mnovafolder, sampleid):
    """
    :param path_to_mnovafolder: path to the folder that contains the Mnova result folders..
    :param sampleid: A string
    :return: pandas DataFrame contains intercept and slope that are calculated based on the std data.
    It can be obtained as std_info['Values']['Intercept']
    """
    std_info_path = path_to_mnovafolder + '\\' + sampleid + r' STD\std_info.xlsx'
    std_info = pd.read_excel(std_info_path, index_col=0, header=0)
    return std_info


def check_consecutive(num_lst):
    """
    Check if the list of numbers are consecutive.
    :param num_lst: A list of numbers.
    :return: Boolean
    """
    return sorted(num_lst) == list(range(min(num_lst), max(num_lst) + 1))


def peak_predictions(peaks_df):
    """
    This makes the predictions for the UV and CAD peak profiles. If it requires the attention,
    then it indicates as 'Check' on the final DataFrame. The images are saved with the predictions.
    :param peaks_df: pd.DataFrame with Well, cmp ID, UV images, and CAD images. Well is the index.
    :return: pd.DataFrame which contains predictions for UV and CAD images. The index is cmp ID.
    """
    # Set the cmp ID column to be index.
    peaks_df = peaks_df.set_index('G#')

    # create the metadata for the prediction.
    MetadataCatalog.get("meta").set(thing_classes=["Bad", "Good"],
                                    thing_colors=[(250, 128, 114), (127, 255, 212)])

    # Build the pre-trained model.
    cfg = get_cfg()
    cfg.merge_from_file(model_zoo.get_config_file("COCO-Detection/faster_rcnn_R_50_FPN_3x.yaml"))
    cfg.MODEL.ROI_HEADS.SCORE_THRESH_TEST = 0.8  # Set threshold for this model
    cfg.MODEL.WEIGHTS = model_path  # Most recently updated model
    cfg.MODEL.DEVICE = 'cpu'  # Run on CPU
    cfg.MODEL.ROI_HEADS.NUM_CLASSES = 2  # Classifications: Bad, Good
    predictor = DefaultPredictor(cfg)

    # Make the predictions.
    column_names = ['UVpeaks', 'CADpeaks']
    prediction_columns = ['UV_predictions', 'CAD_predictions']
    for i, column_name in enumerate(column_names):
        # Create the empty df.column to store the predictions.
        peaks_df[prediction_columns[i]] = ''
        for j, im_array in enumerate(peaks_df[column_name]):
            outputs = predictor(im_array)  # Prediction made
            instances = outputs["instances"]  # Result
            # If len(instances) > 0, the model made some predictions.
            if len(instances) > 0:
                # Get the result with the highest confidence level.
                highest_score_instances = instances[instances.scores == instances.scores.cpu().data.numpy().max()]
                # Output text is created depending on the result.
                # If the peak is predicted to be normal, the output text is '.'.
                if int(highest_score_instances.pred_classes[0].to("cpu").numpy()) == 1:
                    peaks_df.loc[peaks_df.index[j], prediction_columns[i]] = '.'
                # If the peak is predicted to be abnormal, the output text is 'Abnormal peak'.
                elif int(highest_score_instances.pred_classes[0].to("cpu").numpy()) == 0:
                    peaks_df.loc[peaks_df.index[j], prediction_columns[i]] = 'Abnormal peak'
                # If the peak prediction is a value not pre-defined, the output text is 'Undefined prediction'.
                else:
                    peaks_df.loc[peaks_df.index[j], prediction_columns[i]] = 'Undefined prediction'
            # If len(instances) == 0, the prediction was not made.
            else:
                peaks_df.loc[peaks_df.index[j], prediction_columns[i]] = 'No prediction'
    # Remove the unnecessary three columns.
    predictions_df = peaks_df.drop(['Well', 'UVpeaks', 'CADpeaks'], axis=1)
    # Reset the index to keep cmp ID in the dataframe.
    predictions_df = predictions_df.reset_index()
    return predictions_df


def data_process(path_to_mnovadata, path_to_barcode, sampleid, plateids):
    """
    Calculate concentrations and purities, retention time difference among replicates,
    difference between MS - CAD and MS - UV among replicates, and standard deviations for all the data.
    The data is examined for the threshold values and if it doesn't pass the criteria, the comment is added.
    The result is saved as an excel file into the Mnova data folder.
    The FinalResult sheet is formatted to be submission ready. The threshold values are saved in
    X:\analytical\PhysChem Properties\QC\Data Process Templates\thresholds_for_Mnovadataprocess.xlsx
    :param path_to_mnovadata: path to the folder that contains the Mnova result folders.
    :param path_to_barcode: (pd.DataFrame) Path to the barcode information.
    :param sampleid: (string) Sample ID such as WK## or CD##.
    :param plateids: (list) List of plate IDs.
    :return: None
    """
    # Model validation.
    val_path = path_to_mnovadata + '\\' + 'Validation'
    # Get the validation information.
    val_data = pd.read_excel(path_to_mnovadata + r'\Validation\Validation.xlsx', header=0, index_col=0)
    # Create the dataframe to store data.
    val_data_structure = pd.DataFrame({'Well': np.array(val_data.index), 'G#': np.array(val_data.index)})
    # Crop the images
    val_peaks_df = crop_UVCAD_plots.crop_uvcad_plots(val_path, val_data_structure)
    # Make predictions on validation images.
    val_predictions_df = peak_predictions(val_peaks_df)
    # Set the cmp ID column to be index
    val_predictions_df.set_index(['G#'], inplace=True)
    # Assign two empty columns where the validation result will be saved.
    val_predictions_df['UV_check'] = ''
    val_predictions_df['CAD_check'] = ''
    # Make a validation.
    for i, compound_id in enumerate(val_predictions_df.index):
        for j, d_type in enumerate(['UV', 'CAD']):
            # Concatenate the human judge (Bad or Good) and prediction(Abnormal peak or .).
            letter_length = len(val_data.loc[compound_id, d_type] +
                                val_predictions_df.loc[compound_id, d_type + '_predictions'])
            # The prediction is correct if it's "BADAbnormal peak" or "Good."; 16 or 5 letters.
            if letter_length == 5 or letter_length == 16:
                val_predictions_df.loc[compound_id, d_type + '_check'] = ''
            # The prediction is incorrect if it's "BAD." or "GoodAbnormal peak; 4 or 17 letters"
            elif letter_length == 4 or letter_length == 17:
                val_predictions_df.loc[compound_id, d_type + '_check'] = 'Incorrect prediction'
            # Something went wrong if there is other values than the ones stated above.
            else:
                val_predictions_df.loc[compound_id, d_type + '_check'] = 'Error'
    # Extract only the necessary information to output.
    val_df = val_predictions_df[['UV_check', 'CAD_check']]

    # Get threshold information from thresholds_info_path
    thresholds_info = pd.read_excel(thresholds_info_path, usecols=[0, 1], header=0, index_col=0)
    conc_stdev_threshold = thresholds_info.loc[('conc_stdev_threshold', 'Values')]
    purity_stdev_threshold = thresholds_info.loc[('purity_stdev_threshold', 'Values')]
    matched_rt_threshold = thresholds_info.loc[('matched_RT_threshold', 'Values')]
    diff_uv_maxthreshold = thresholds_info.loc[('diff_UV_maxthreshold', 'Values')]
    diff_cad_maxthreshold = thresholds_info.loc[('diff_CAD_maxthreshold', 'Values')]
    diff_uv_minthreshold = thresholds_info.loc[('diff_UV_minthreshold', 'Values')]
    diff_cad_minthreshold = thresholds_info.loc[('diff_CAD_minthreshold', 'Values')]
    min_purity = thresholds_info.loc[('min_purity', 'Values')]
    min_conc = thresholds_info.loc[('min_conc', 'Values')]

    # Sample type is either Weekly including special request or Chem Div.
    sample_type = sampleid[0:2]

    # Get slope and intercept from the latest std validation.
    std_info = get_std_info(path_to_mnovadata, sampleid=sampleid)
    slope = std_info.loc[('Slope', 'Values')]
    intercept = std_info.loc[('Intercept', 'Values')]

    # Create the frameworks to save the data of all plate IDs..
    data_dict = {}
    cad_areas_dict = {}
    purities_dict = {}
    matched_rts_dict = {}
    diff_uvs_dict = {}
    diff_cads_dict = {}
    uv_peak_predictions_dict = {}
    cad_peak_predictions_dict = {}
    result_folders = []

    # Get the folder names of all Mnova results.
    for folder_name in os.listdir(path_to_mnovadata):
        if 'STD' not in folder_name and 'Validation' not in folder_name:
            result_folders.append(folder_name)

    # Get the necessary information from the Mnova results.
    count = 0
    for plateid in plateids:
        # Create the data dictionary for each plate ID.
        matched_rts = {}
        diff_uvs = {}
        diff_cads = {}
        cad_areas = {}
        purities = {}
        uv_peak_predictions = {}
        cad_peak_predictions = {}

        # Loop through the replicates for the plate ID.
        for i, rep in enumerate(replicates):
            # Get the Mnova data.
            mnova_result_path = path_to_mnovadata + '\\' + result_folders[i + count]
            mnova_data = pd.read_csv(mnova_result_path + r'\MSQCResults.csv', index_col=4, header=0)
            mnova_data.index.name = 'G#'
            mnova_data.insert(loc=4, column='GNumber', value=mnova_data.index)
            data_dict[plateid + '-' + rep] = mnova_data

            # Data structure for peak images.
            peaks_datastructure = mnova_data[['Well', 'GNumber']]
            # Crop the UV and CAD plots.
            peaks_df = crop_UVCAD_plots.crop_uvcad_plots(mnova_result_path + r'\pdf', peaks_datastructure)
            peaks_df.insert(loc=0, column='G#', value=peaks_df['GNumber'])
            # Check for any abnormal peaks.
            peak_predictions_df = peak_predictions(peaks_df)
            peak_predictions_df.set_index('G#', inplace=True)

            uv_peak_predictions[rep] = peak_predictions_df['UV_predictions'].to_frame(name=rep)
            cad_peak_predictions[rep] = peak_predictions_df['CAD_predictions'].to_frame(name=rep)

            # Extract the data needed.
            cad_areas[rep] = mnova_data['Conc (mM)'].to_frame(name=rep)
            purities[rep] = mnova_data['Purity (%)'].to_frame(name=rep)
            matched_rts[rep] = mnova_data['Matched RT'].to_frame(name=rep)
            diff_uvs[rep] = mnova_data['Diff PDA'].to_frame(name=rep)
            diff_cads[rep] = mnova_data['Diff CAD'].to_frame(name=rep)
        count += len(replicates)  # Increase by the number of replicates for each loop.

        # Horizontally concatenate the replicates' data.
        uv_peak_predictions_dict[plateid] = uv_peak_predictions[replicates[0]].merge(
            uv_peak_predictions[replicates[1]],
            how='left', on='G#')
        uv_peak_predictions_dict[plateid] = uv_peak_predictions_dict[plateid].merge(
            uv_peak_predictions[replicates[2]],
            how='left', on='G#')
        cad_peak_predictions_dict[plateid] = cad_peak_predictions[replicates[0]].merge(
            cad_peak_predictions[replicates[1]], how='left', on='G#')
        cad_peak_predictions_dict[plateid] = cad_peak_predictions_dict[plateid].merge(
            cad_peak_predictions[replicates[2]], how='left', on='G#')

        matched_rts_dict[plateid] = matched_rts[replicates[0]].merge(matched_rts[replicates[1]], how='left',
                                                                     on='G#')
        matched_rts_dict[plateid] = matched_rts_dict[plateid].merge(matched_rts[replicates[2]], how='left',
                                                                    on='G#')

        cad_areas_dict[plateid] = cad_areas[replicates[0]].merge(cad_areas[replicates[1]], how='left', on='G#')
        cad_areas_dict[plateid] = cad_areas_dict[plateid].merge(cad_areas[replicates[2]], how='left', on='G#')

        purities_dict[plateid] = purities[replicates[0]].merge(purities[replicates[1]], how='left', on='G#')
        purities_dict[plateid] = purities_dict[plateid].merge(purities[replicates[2]], how='left', on='G#')

        diff_uvs_dict[plateid] = diff_uvs[replicates[0]].merge(diff_uvs[replicates[1]], how='left', on='G#')
        diff_uvs_dict[plateid] = diff_uvs_dict[plateid].merge(diff_uvs[replicates[2]], how='left', on='G#')

        diff_cads_dict[plateid] = diff_cads[replicates[0]].merge(diff_cads[replicates[1]], how='left', on='G#')
        diff_cads_dict[plateid] = diff_cads_dict[plateid].merge(diff_cads[replicates[2]], how='left', on='G#')

    # Get the barcode info [Gnumber, Molecular Masses, Project, Chemist]
    col = [4, 9, 12, 13]
    barcode_df = pd.read_excel(path_to_barcode, header=0, index_col=0, usecols=col)
    # barcode_df.set_index('GNumber', inplace=True)
    # Separate the barcode information into each plate ID.
    barcode_dict = {}
    for i, j in enumerate(plateids):
        barcode_dict[j] = barcode_df.iloc[i * platesize: (i + 1) * platesize]

    # Calculate the concentrations for each replicate.
    for plateid in plateids:
        for i, rep in enumerate(replicates):
            for compound_id, area in cad_areas_dict[plateid][rep].iteritems():
                mm = barcode_dict[plateid].loc[compound_id][0]  # Molecular mass
                # If the values are not given, ignore them.
                if area == '-' or area == 'not matched':
                    pass
                else:
                    conc = (float(area) - intercept) / slope * 100 / mm
                    # If the concentration is calculated to be negative, then the concentration is set to be 0.
                    if conc < 0:
                        cad_areas_dict[plateid].loc[(compound_id, rep)] = 0
                    # Concentration values are rounded to one decimal place.
                    else:
                        cad_areas_dict[plateid].loc[(compound_id, rep)] = round(conc, 1)

    # Concentration data process.
    for plateid in plateids:
        cad_areas_dict[plateid].loc[:, 'Average Conc (mM)'] = ''
        cad_areas_dict[plateid].loc[:, 'STDEV (mM)'] = ''
        cad_areas_dict[plateid].loc[:, 'STDEV check'] = ''
        cad_areas_dict[plateid].loc[:, 'No values for rep'] = ''
        for compound_id in cad_areas_dict[plateid].index:
            replicates_idx = list(range(0, len(replicates)))
            # Get the index of values that are not numerical.
            idx_no_values = [i for i, j in
                             enumerate(list(cad_areas_dict[plateid].loc[compound_id][0:len(replicates)]))
                             if not isinstance(j, numbers.Number)]
            idx_no_quant = [i for i, j in
                            enumerate(list(cad_areas_dict[plateid].loc[compound_id][0:len(replicates)]))
                            if j == '-']
            idx_no_match = [i for i, j in
                            enumerate(list(cad_areas_dict[plateid].loc[compound_id][0:len(replicates)]))
                            if j == 'not matched']
            # Get the index of values that are numerical.
            for i in idx_no_values:
                replicates_idx.remove(i)
            # Get the concentrations
            concentrations = []
            for i in replicates_idx:
                concentrations.append(cad_areas_dict[plateid].loc[compound_id][i])
            # If there is more than one "-" out of all replicates, then the CAD is not quantified.
            if len(idx_no_quant) > 1:
                cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')] = 'No CAD quantification  '
            # If there is more than one "not matched" out of all replicates, then no target mass was found.
            elif len(idx_no_match) > 1:
                cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')] = 'No target mass  '
            # Calculate the average and standard deviation of the concentrations.
            else:
                cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')] = round(np.mean(concentrations), 1)
                cad_areas_dict[plateid].loc[(compound_id, 'STDEV (mM)')] = round(np.std(concentrations), 0)
            # Make comments to the values.
            if np.std(concentrations) > conc_stdev_threshold:
                cad_areas_dict[plateid].loc[(compound_id, 'STDEV check')] = 'Check'
            if len(idx_no_values) != 0:
                cad_areas_dict[plateid].loc[(compound_id, 'No values for rep')] = str(np.array(idx_no_values) + 1)

    # Purity data process.
    for plateid in plateids:
        purities_dict[plateid].loc[:, 'Average Purity (%)'] = ''
        purities_dict[plateid].loc[:, 'STDEV (%)'] = ''
        purities_dict[plateid].loc[:, 'STDEV check'] = ''
        purities_dict[plateid].loc[:, 'No values for rep'] = ''
        # Get the index of purity values that are not numerical.
        for compound_id in purities_dict[plateid].index:
            idx_no_values = [i for i, j in
                             enumerate(list(purities_dict[plateid].loc[compound_id][0:len(replicates)]))
                             if j == '-' or j == 'not matched']
            idx_no_quant = [i for i, j in
                            enumerate(list(purities_dict[plateid].loc[compound_id][0:len(replicates)]))
                            if j == '-']
            idx_no_match = [i for i, j in
                            enumerate(list(purities_dict[plateid].loc[compound_id][0:len(replicates)]))
                            if j == 'not matched']
            # Get the index of purity values that are numerical.
            replicates_idx = list(range(0, len(replicates)))
            for i in idx_no_values:
                replicates_idx.remove(i)
            # Get the purity values.
            purities = []
            for r in replicates_idx:
                purities.append(float(purities_dict[plateid].loc[compound_id][r]))
            # If there is more than one "-" out of all replicates, then the sample is no quantified.
            if len(idx_no_quant) > 1:
                purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')] = 'No UV quantification'
            # If there is more than one "not matched" out of all replicates, then no target mass was found.
            elif len(idx_no_match) > 1:
                purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')] = 'No target mass'
            # Calculate the average and standard deviation of the purities.
            else:
                purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')] = round(np.mean(purities), 0)
                purities_dict[plateid].loc[(compound_id, 'STDEV (%)')] = round(np.std(purities), 0)
            # Make comments to the values.
            if np.std(purities) > purity_stdev_threshold:
                purities_dict[plateid].loc[(compound_id, 'STDEV check')] = 'Check'
            if len(idx_no_values) != 0:
                purities_dict[plateid].loc[(compound_id, 'No values for rep')] = str(np.array(idx_no_values) + 1)

    # Check matched RT.
    for plateid in plateids:
        matched_rts_dict[plateid].loc[:, 'STDEV'] = ''
        matched_rts_dict[plateid].loc[:, 'Large STDEV'] = ''
        # Get the index of purity values that are not numerical.
        for compound_id in cad_areas_dict[plateid].index:
            # Get the index of matched retention time that are not numerical.
            idx_no_values = [i for i, j in enumerate(list(matched_rts_dict[plateid].loc[compound_id][0:3])) if
                             j == '-']
            # Get the index of matched retention time that are numerical.
            replicates_idx = list(range(0, len(replicates)))
            for i in idx_no_values:
                replicates_idx.remove(i)
            # Get the retention times.
            rts = []
            for r in replicates_idx:
                rts.append(float(matched_rts_dict[plateid].loc[compound_id][r]))
            # Calculate the standard deviation of retention times.
            matched_rts_dict[plateid].loc[(compound_id, 'STDEV')] = round(np.std(rts), 1)
            # Make comments.
            if np.std(rts) > matched_rt_threshold:
                matched_rts_dict[plateid].loc[(compound_id, 'Large STDEV')] = 'Check'

    # Check the retention time difference for MS-UV and MS-CAD.
    for plateid in plateids:
        diff_uvs_dict[plateid].loc[:, 'MS-UV time lag out of range for rep:'] = ''
        diff_cads_dict[plateid].loc[:, 'MS-CAD time lag out of range for rep:'] = ''

        for compound_id in diff_uvs_dict[plateid].index:
            # Get the index of retention time that are not numerical for each UV and CAD.
            uv_idx_no_values = [i for i, j in
                                enumerate(list(diff_uvs_dict[plateid].loc[compound_id][0:len(replicates)]))
                                if j == '-']
            cad_idx_no_values = [i for i, j in
                                 enumerate(list(diff_cads_dict[plateid].loc[compound_id][0:len(replicates)]))
                                 if j == '-']
            # Get the index of retention time that are numerical for each UV and CAD.
            uv_replicates_idx = list(range(0, len(replicates)))
            cad_replicates_idx = list(range(0, len(replicates)))
            for i in uv_idx_no_values:
                uv_replicates_idx.remove(i)
            for i in cad_idx_no_values:
                cad_replicates_idx.remove(i)
            # Pick up the replicate numbers of the values that are out of thresholds set in the thresholds_info_path.
            check_uvs = []
            check_cads = []
            for r in uv_replicates_idx:
                if float(diff_uvs_dict[plateid].loc[compound_id][r]) > diff_uv_maxthreshold or float(
                        diff_uvs_dict[plateid].loc[compound_id][r]) < diff_uv_minthreshold:
                    check_uvs.append(r)
            for r in cad_replicates_idx:
                if float(diff_cads_dict[plateid].loc[compound_id][r]) > diff_cad_maxthreshold or float(
                        diff_cads_dict[plateid].loc[compound_id][r]) < diff_cad_minthreshold:
                    check_cads.append(r)
            # Make comments.
            if len(check_uvs) != 0:
                diff_uvs_dict[plateid].loc[
                    (compound_id, 'MS-UV time lag out of range for rep:')] = str(np.array(check_uvs) + 1)
            elif len(check_cads) != 0:
                diff_cads_dict[plateid].loc[
                    (compound_id, 'MS-CAD time lag out of range for rep:')] = str(np.array(check_cads) + 1)
            else:
                pass

    # Check the predictions.
    data_types = ['UV', 'CAD']
    for i, d in enumerate([uv_peak_predictions_dict, cad_peak_predictions_dict]):
        label = data_types[i] + ' Check the rep:'
        for plateid in plateids:
            # Set a column for the prediction comment.
            d[plateid][label] = ''
            # Get the index of values that are 'Abnormal peak'.
            for compound_id in d[plateid].index:
                idx_abnormal_peak = [i + 1 for i, j in
                                     enumerate(list(d[plateid].loc[compound_id][0:len(replicates)])) if
                                     j == 'Abnormal peak']
                # If there are more than one abnormal peak out of the replicates,
                # leave the replicate numbers on the comment.
                if len(idx_abnormal_peak) > 1:
                    d[plateid].loc[(compound_id, label)] = idx_abnormal_peak

    # Combine UV and CAD comments into one.
    for plateid in plateids:
        # Check if the index columns of UV and CAD dataframe match.
        if (cad_peak_predictions_dict[plateid].index == uv_peak_predictions_dict[plateid].index).all():
            uv_peak_predictions_dict[plateid]['CAD Check the rep:'] = cad_peak_predictions_dict[plateid][
                'CAD Check the rep:']
        else:
            ValueError('G number does not match.')

    # Remove unnecessary rep# columns.
    for p in plateids:
        uv_peak_predictions_dict[p].pop('rep1')
        uv_peak_predictions_dict[p].pop('rep2')
        uv_peak_predictions_dict[p].pop('rep3')

    # Summarize the average of purity and concentration data using the existing dataframe
    # (rep1 of initial Mnova dataframe).
    for plateid in plateids:
        for compound_id in data_dict[plateid + '-rep1'].index:
            # Add comments.
            if isinstance(cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')], numbers.Number):
                if cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')] < min_conc:
                    conc_analyst_comment = 'Low concentration  '
                else:
                    conc_analyst_comment = ''
            elif isinstance(cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')], str):
                conc_analyst_comment = cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')]
                cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')] = 0
            else:
                conc_analyst_comment = ''
            if isinstance(purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')], numbers.Number):
                if purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')] < min_purity:
                    purity_analyst_comment = 'Low purity'
                else:
                    purity_analyst_comment = ''
            elif isinstance(purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')], str):
                purity_analyst_comment = purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')]
                purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')] = 0
            else:
                purity_analyst_comment = ''
            data_dict[plateid + '-rep1'].loc[(compound_id, 'Analyst Comments')] = conc_analyst_comment + \
                                                                                  purity_analyst_comment
            # Update the columns, Well, Conc, and Purity, with the values determined above.
            data_dict[plateid + '-rep1'].loc[(compound_id, 'Well')] = \
                sampleid + plateid + data_dict[plateid + '-rep1'].loc[(compound_id, 'Well')][-4:]
            data_dict[plateid + '-rep1'].loc[(compound_id, 'Conc (mM)')] = \
                cad_areas_dict[plateid].loc[(compound_id, 'Average Conc (mM)')]
            data_dict[plateid + '-rep1'].loc[(compound_id, 'Purity (%)')] = \
                purities_dict[plateid].loc[(compound_id, 'Average Purity (%)')]
    # Create the framework for the final result.
    df_list = [data_dict[plateid + '-rep1'] for plateid in plateids]
    # Vertically concatenate the df in the list.
    finalresult_df = pd.concat(df_list, axis=0)
    # Remove the columns that are not needed on the final data report.
    finalresult_df = finalresult_df.drop(['Full Formula', 'Matched RT', 'Diff PDA', 'Diff CAD'], axis=1)
    # Make a different style of report depending on the sample type, WK or CD.
    if sample_type == 'WK':  # For the weekly or special request samples.
        # Create a DataFrame for the emailing format.
        email_format_df = pd.DataFrame({'Num': list(range(1, len(finalresult_df) + 1)),
                                        'G-Number': finalresult_df['GNumber'],
                                        'Project': barcode_df['Project'],
                                        'Exp Conc in DMSO (mM)': finalresult_df['Expected Conc'],
                                        'Obs conc in DMSO (mM)': finalresult_df['Conc (mM)'],
                                        'Purity (UV) in DMSO': finalresult_df['Purity (%)'],
                                        'Comments': finalresult_df['Analyst Comments']
                                        }, index=finalresult_df['GNumber']
                                       )
        # Make sure the data from finalresult_df and barcode_df match.
        for compound_id in email_format_df.index:
            email_format_df.loc[(compound_id, 'Project')] = barcode_df.loc[(compound_id, 'Project')]
        data_framework = pd.DataFrame({'Well': finalresult_df['Well'], 'GNumber': finalresult_df['GNumber']},
                                      index=finalresult_df['GNumber'])

    elif sample_type == 'CD':  # For the chem div samples.
        # Create the DataFrame for Chem Div data loading.
        finalresult_df = pd.DataFrame({'Num': list(range(1, len(finalresult_df) + 1)),
                                       'LCMS file name': finalresult_df['Well'],
                                       'G-num/SMDI-num': finalresult_df['GNumber'],
                                       'SMDI ID': finalresult_df['SMDI ID'],
                                       'Target MW': finalresult_df['Observed Mass'],
                                       'UV(254)% Purity': finalresult_df['Purity (%)'],
                                       'Ob. Conc (mM)': finalresult_df['Conc (mM)'],
                                       'Project': barcode_df['Project'],
                                       'Chemist': barcode_df['Chemist'],
                                       'Comment_Conc': finalresult_df['Analyst Comments']},
                                      index=finalresult_df['GNumber'])
        # Make sure the data from finalresult_df and barcode_df match.
        for compound_id in finalresult_df.index:
            finalresult_df.loc[(compound_id, 'Project')] = barcode_df.loc[(compound_id, 'Project')]
            finalresult_df.loc[(compound_id, 'Chemist')] = barcode_df.loc[(compound_id, 'Chemist')]
        # This has the columns, Well, GNumber, and cmp ID, and will be used as a data framework for the final reports.
        data_framework = pd.DataFrame({'Well': finalresult_df['LCMS file name'],
                                       'GNumber': finalresult_df['G-num/SMDI-num']},
                                      index=finalresult_df['G-num/SMDI-num'])
    data_framework.index.name = 'G#'

    # Path where the final result will be saved.
    finalresult_folder = os.path.abspath(os.path.join(path_to_mnovadata, os.pardir))

    # Frameworks for the analysis outputs.
    conc_df = data_framework
    purities_df = data_framework
    matched_rts_df = data_framework
    diff_uvs_df = data_framework
    diff_cads_df = data_framework
    peaks_df = data_framework
    data_dict_list = [cad_areas_dict, purities_dict, matched_rts_dict, diff_uvs_dict, diff_cads_dict,
                      uv_peak_predictions_dict]
    finaldata_df_list = [conc_df, purities_df, matched_rts_df, diff_uvs_df, diff_cads_df, peaks_df]

    # Loop through all the data analyzed above.
    for i, d in enumerate(data_dict_list):
        df_list = [d[plateid] for plateid in plateids]
        # Vertically concatenate the data (measured/calculated data and comments above) of plate IDs.
        df = pd.concat(df_list, axis=0)
        # Horizontally merge the data above into the data framework.
        finaldata_df_list[i] = finaldata_df_list[i].merge(df, how='left', on='G#')
    # Add the validation result to the last of the data list.
    finaldata_df_list.append(val_df)
    # Add the submission sheet as the first sheet on the excel file.
    finaldata_df_list.insert(0, finalresult_df)

    # Define an Excel writer object and the target file. The result will be named as 'sample ID'_Result.xlsx.
    finalresult_excel = finalresult_folder + '\\' + sampleid + "_Result.xlsx"
    excelwriter = pd.ExcelWriter(finalresult_excel, engine="xlsxwriter")
    excelsheet_names = ['FinalData', 'Conc', 'Purity', 'MatchedRT', 'DiffUV', 'DiffCAD', 'PeakCheck',
                        'Validation']

    # If the sample is weekly analysis, add the emailing-format data sheet.
    if sample_type == 'WK':
        finaldata_df_list.append(email_format_df)
        excelsheet_names.append('EmailingFormat')

    # Make all the excel sheets.
    for i, df in enumerate(finaldata_df_list):
        df.to_excel(excelwriter, sheet_name=excelsheet_names[i], index=False)

    # Save the file
    excelwriter.save()

    # Insert the Data Loader path to the final result (for loading the data on SMDI).
    wb = openpyxl.load_workbook(finalresult_excel)
    # Insert a row to the top of the final data sheet.
    finalsheet_sheet = wb['FinalData']
    finalsheet_sheet.insert_rows(1)
    # Data loader path for the weekly sample.
    if sample_type == 'WK':
        finalsheet_sheet.cell(row=1, column=1).value = weeklyanalysis_dataloader_path
    # Data loader path for the chem div sample.
    elif sample_type == 'CD':
        finalsheet_sheet.cell(row=1, column=1).value = chemdiv_dataloader_path
        finalsheet_sheet.cell(row=1, column=2).value = chemdiv_comment
    # Update the excel file.
    wb.save(finalresult_excel)
    return finaldata_df_list

